from django.shortcuts import render, redirect
from django.http import JsonResponse, FileResponse, Http404, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from datetime import datetime, timedelta
import json
import os
from django.core.serializers.json import DjangoJSONEncoder
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage
from django.conf import settings
from io import BytesIO

# Importar los modelos desde partida
from partida.models import (
    Matricula, Categoria, Alumno, Persona, 
    Sesionentrenamiento, Asistencia, CalificacionObjetivos,
    Objetivos, Entrenamiento, EntrenamientoObjetivo  # AGREGAR EntrenamientoObjetivo
)
# Importar el generador de PDF
from .utils.generador_pdf import generar_certificado_individual

from django.contrib.auth.decorators import login_required
from partida.views import prevent_cache 

# Create your views here.
@login_required
@prevent_cache
def index(request):
    """Vista principal de reportes"""
    categorias = Categoria.objects.all().order_by('idcategoria')  # Cambiado de 'nom_categoria' a 'idcategoria'
    
    context = {
        'categorias': categorias,
        'tipoReporte': None,
    }
    
    return render(request, 'reportes.html', context)


@login_required
def generar_reporte_individual(request):
    """Genera reporte individual de un alumno"""
    if request.method == 'POST':
        categoria_id = request.POST.get('categoria')
        matricula_id = request.POST.get('idAlumno')
        fecha_inicio_str = request.POST.get('fechaInicio')
        fecha_fin_str = request.POST.get('fechaFin')
        
        categorias = Categoria.objects.all().order_by('idcategoria')
        
        try:
            # Convertir fechas de string a datetime
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            
            # Ajustar fecha_fin para incluir todo el día
            from datetime import timedelta
            fecha_fin = fecha_fin + timedelta(days=1)
            
            # Obtener la matrícula y sus datos
            matricula = Matricula.objects.select_related(
                'fk_alumno__fk_persona_alumno',
                'fk_categoria'
            ).get(idmatricula=matricula_id)
            
            persona = matricula.fk_alumno.fk_persona_alumno
            nombre_alumno = f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip()
            nombre_categoria = matricula.fk_categoria.nom_categoria
            
            # Obtener todas las asistencias del alumno en el rango de fechas
            asistencias = Asistencia.objects.filter(
                fk_matricula_ms=matricula,
                fecha_asistencia__gte=fecha_inicio,
                fecha_asistencia__lt=fecha_fin
            ).select_related(
                'fk_sesion_ms__fk_entrenamiento'
            ).order_by('fecha_asistencia')
            
            reportes_individuales = []
            
            for asistencia in asistencias:
                # Obtener el entrenamiento de la sesión
                entrenamiento = asistencia.fk_sesion_ms.fk_entrenamiento
                
                # Determinar estado de asistencia
                estado_asistencia = 'PRESENTE' if asistencia.asistencia == 1 else 'AUSENTE'
                
                # Obtener calificaciones de objetivos SOLO si objetivo_evaluado = True (1)
                calificaciones = CalificacionObjetivos.objects.filter(
                    fk_asistencia=asistencia,
                    objetivo_evaluado=True  # Solo objetivos que fueron evaluados
                ).select_related('fk_asistencia')
                
                # Si hay calificaciones, procesar cada una
                if calificaciones.exists():
                    for calificacion in calificaciones:
                        # Obtener el nombre del objetivo
                        try:
                            objetivo = Objetivos.objects.get(idobjetivos=calificacion.id_objetivo)
                            nombre_objetivo = objetivo.nom_objetivo
                        except Objetivos.DoesNotExist:
                            nombre_objetivo = f"Objetivo ID: {calificacion.id_objetivo}"
                        
                        # Determinar si el objetivo fue cumplido
                        objetivo_cumplido = 'SÍ' if calificacion.evaluacion else 'NO'
                        
                        reportes_individuales.append({
                            'fechaEntrenamiento': asistencia.fecha_asistencia.strftime('%d/%m/%Y %H:%M'),
                            'asistencia': estado_asistencia,
                            'tipoEntrenamiento': entrenamiento.nom_entrenamiento,
                            'nombreObjetivo': nombre_objetivo,
                            'objetivoCumplido': objetivo_cumplido,
                            'observaciones': calificacion.observaciones if calificacion.observaciones else 'Sin observaciones'
                        })
                else:
                    # Si no hay calificaciones, mostrar solo la asistencia
                    reportes_individuales.append({
                        'fechaEntrenamiento': asistencia.fecha_asistencia.strftime('%d/%m/%Y %H:%M'),
                        'asistencia': estado_asistencia,
                        'tipoEntrenamiento': entrenamiento.nom_entrenamiento,
                        'nombreObjetivo': 'Sin objetivos evaluados',
                        'objetivoCumplido': 'N/A',
                        'observaciones': asistencia.observaciones if asistencia.observaciones else 'Sin observaciones'
                    })
            
            context = {
                'categorias': categorias,
                'tipoReporte': 'individual',
                'reportesIndividuales': reportes_individuales,
                'nombreCategoria': nombre_categoria,
                'nombreAlumno': nombre_alumno,
                'categoriaId': categoria_id,
                'matriculaId': matricula_id,  # AGREGAR ESTA LÍNEA
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
            }
            
        except Matricula.DoesNotExist:
            context = {
                'categorias': categorias,
                'tipoReporte': 'individual',
                'reportesIndividuales': [],
                'nombreCategoria': '',
                'nombreAlumno': '',
                'categoriaId': categoria_id,
                'matriculaId': matricula_id,  # AGREGAR ESTA LÍNEA
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
                'error': 'No se encontró la matrícula seleccionada'
            }
        except Exception as e:
            context = {
                'categorias': categorias,
                'tipoReporte': 'individual',
                'reportesIndividuales': [],
                'nombreCategoria': '',
                'nombreAlumno': '',
                'categoriaId': categoria_id,
                'matriculaId': matricula_id,  # AGREGAR ESTA LÍNEA
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
                'error': f'Error al generar el reporte: {str(e)}'
            }
        
        return render(request, 'reportes.html', context)
    
    return redirect('reportes:index')


@login_required
def generar_reporte_grupal(request):
    """Genera reporte grupal de una categoría"""
    if request.method == 'POST':
        categoria_id = request.POST.get('categoria')
        fecha_inicio_str = request.POST.get('fechaInicio')
        fecha_fin_str = request.POST.get('fechaFin')
        
        categorias = Categoria.objects.all().order_by('idcategoria')  # Cambiado de 'nom_categoria' a 'idcategoria'
        
        try:
            # Convertir fechas
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            
            # Ajustar fecha_fin para incluir todo el día
            from datetime import timedelta
            fecha_fin = fecha_fin + timedelta(days=1)
            
            # Obtener información de la categoría
            categoria = Categoria.objects.get(idcategoria=categoria_id)
            nombre_categoria = categoria.nom_categoria
            
            # Obtener todas las matrículas activas de la categoría ORDENADAS por apellidos
            matriculas = Matricula.objects.filter(
                fk_categoria_id=categoria_id,
                estado_matricula=1
            ).select_related('fk_alumno__fk_persona_alumno').order_by(
                'fk_alumno__fk_persona_alumno__ape1_persona',
                'fk_alumno__fk_persona_alumno__ape2_persona',
                'fk_alumno__fk_persona_alumno__nom1_persona',
                'fk_alumno__fk_persona_alumno__nom2_persona'
            )
            
            reportes_grupales = []
            suma_asistencia = 0
            suma_objetivos = 0
            
            for matricula in matriculas:
                persona = matricula.fk_alumno.fk_persona_alumno
                
                # Formatear nombre: Apellido1 Apellido2, Nombre1 Nombre2
                apellidos = f"{persona.ape1_persona} {persona.ape2_persona or ''}".strip()
                nombres = f"{persona.nom1_persona} {persona.nom2_persona or ''}".strip()
                nombre_completo = f"{apellidos}, {nombres}"
                
                # Obtener todas las asistencias del alumno en el rango de fechas
                asistencias = Asistencia.objects.filter(
                    fk_matricula_ms=matricula,
                    fecha_asistencia__gte=fecha_inicio,
                    fecha_asistencia__lt=fecha_fin
                ).select_related('fk_sesion_ms')
                
                total_entrenamientos = asistencias.count()
                
                if total_entrenamientos == 0:
                    continue  # Saltar alumnos sin entrenamientos en el período
                
                # Contar asistencias (presente = 1)
                total_asistencias = asistencias.filter(asistencia=1).count()
                
                # Calcular porcentaje de asistencia
                porcentaje_asistencia = (total_asistencias / total_entrenamientos * 100) if total_entrenamientos > 0 else 0
                
                # Obtener todas las calificaciones de objetivos evaluados
                calificaciones = CalificacionObjetivos.objects.filter(
                    fk_asistencia__in=asistencias,
                    objetivo_evaluado=True  # Solo objetivos evaluados
                )
                
                total_objetivos_evaluados = calificaciones.count()
                
                # Contar objetivos cumplidos (evaluacion = True)
                total_objetivos_cumplidos = calificaciones.filter(evaluacion=True).count()
                
                # Calcular porcentaje de objetivos cumplidos
                porcentaje_objetivos = (total_objetivos_cumplidos / total_objetivos_evaluados * 100) if total_objetivos_evaluados > 0 else 0
                
                reportes_grupales.append({
                    'matriculaId': matricula.idmatricula,  # NUEVO: ID de matrícula
                    'nombreCompleto': nombre_completo,
                    'totalEntrenamientos': total_entrenamientos,
                    'totalAsistencias': total_asistencias,
                    'porcentajeAsistencia': round(porcentaje_asistencia, 1),
                    'totalObjetivosEvaluados': total_objetivos_evaluados,
                    'totalObjetivosCumplidos': total_objetivos_cumplidos,
                    'porcentajeObjetivos': round(porcentaje_objetivos, 1)
                })
                
                suma_asistencia += porcentaje_asistencia
                suma_objetivos += porcentaje_objetivos
            
            # Calcular promedios generales
            total_alumnos = len(reportes_grupales)
            promedio_asistencia = (suma_asistencia / total_alumnos) if total_alumnos > 0 else 0
            promedio_objetivos = (suma_objetivos / total_alumnos) if total_alumnos > 0 else 0
            
            context = {
                'categorias': categorias,
                'tipoReporte': 'grupal',
                'reportesGrupales': reportes_grupales,
                'reportesGrupalesJSON': json.dumps(reportes_grupales, cls=DjangoJSONEncoder),  # AGREGAR ESTA LÍNEA
                'nombreCategoria': nombre_categoria,
                'categoriaId': categoria_id,
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
                'sumaAsistencia': round(promedio_asistencia, 1),
                'sumaObjetivos': round(promedio_objetivos, 1),
                'totalAlumnos': total_alumnos,
            }
            
        except Categoria.DoesNotExist:
            context = {
                'categorias': categorias,
                'tipoReporte': 'grupal',
                'reportesGrupales': [],
                'nombreCategoria': '',
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
                'sumaAsistencia': 0,
                'sumaObjetivos': 0,
                'error': 'No se encontró la categoría seleccionada'
            }
        except Exception as e:
            context = {
                'categorias': categorias,
                'tipoReporte': 'grupal',
                'reportesGrupales': [],
                'nombreCategoria': '',
                'fechaInicio': fecha_inicio_str,
                'fechaFin': fecha_fin_str,
                'sumaAsistencia': 0,
                'sumaObjetivos': 0,
                'error': f'Error al generar el reporte: {str(e)}'
            }
        
        return render(request, 'reportes.html', context)
    
    return redirect('reportes:index')


@login_required
@require_http_methods(["GET"])
def obtener_alumnos_por_categoria(request):
    """API para obtener alumnos por categoría"""
    categoria_id = request.GET.get('categoriaId')
    
    if not categoria_id:
        return JsonResponse({'success': False, 'error': 'Categoría no especificada'})
    
    try:
        # Obtener matrículas activas de la categoría seleccionada
        matriculas = Matricula.objects.filter(
            fk_categoria_id=categoria_id,
            estado_matricula=1  # Solo matrículas activas
        ).select_related('fk_alumno__fk_persona_alumno').order_by(
            'fk_alumno__fk_persona_alumno__ape1_persona',
            'fk_alumno__fk_persona_alumno__nom1_persona'
        )
        
        alumnos_data = []
        for matricula in matriculas:
            persona = matricula.fk_alumno.fk_persona_alumno
            nombre_completo = f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip()
            alumnos_data.append({
                'idmatricula': matricula.idmatricula,
                'nombre_completo': nombre_completo
            })
        
        return JsonResponse({
            'success': True,
            'alumnos': alumnos_data
        })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def generar_certificado_pdf(request, matricula_id):
    """Genera y descarga el certificado PDF para un alumno"""
    if request.method == 'GET':
        fecha_inicio_str = request.GET.get('fechaInicio')
        fecha_fin_str = request.GET.get('fechaFin')
        
        try:
            # Convertir fechas
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            fecha_fin_ajustada = fecha_fin + timedelta(days=1)
            
            # Obtener la matrícula y datos del alumno
            matricula = Matricula.objects.select_related(
                'fk_alumno__fk_persona_alumno',
                'fk_categoria'
            ).get(idmatricula=matricula_id)
            
            persona = matricula.fk_alumno.fk_persona_alumno
            
            # Preparar datos del alumno
            datos_alumno = {
                'nombre_completo': f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip(),
                'tipo_identificacion': persona.tipo_identidad,
                'numero_identificacion': persona.id_persona,
                'id_persona': persona.id_persona,
                'categoria': matricula.fk_categoria.nom_categoria,
            }
            
            # Obtener estadísticas del reporte
            asistencias = Asistencia.objects.filter(
                fk_matricula_ms=matricula,
                fecha_asistencia__gte=fecha_inicio,
                fecha_asistencia__lt=fecha_fin_ajustada
            ).order_by('fecha_asistencia')  # ORDENAR por fecha para obtener la primera
            
            total_entrenamientos = asistencias.count()
            total_asistencias = asistencias.filter(asistencia=1).count()
            porcentaje_asistencia = (total_asistencias / total_entrenamientos * 100) if total_entrenamientos > 0 else 0
            
            # NUEVO: Obtener la primera fecha de asistencia real
            primera_fecha_real = None
            if asistencias.exists():
                primera_fecha_real = asistencias.first().fecha_asistencia
            
            # Obtener calificaciones de objetivos
            calificaciones = CalificacionObjetivos.objects.filter(
                fk_asistencia__in=asistencias,
                objetivo_evaluado=True
            )
            
            total_objetivos_evaluados = calificaciones.count()
            total_objetivos_cumplidos = calificaciones.filter(evaluacion=True).count()
            porcentaje_objetivos = (total_objetivos_cumplidos / total_objetivos_evaluados * 100) if total_objetivos_evaluados > 0 else 0
            
            # Preparar datos del reporte
            datos_reporte = {
                'fecha_inicio': primera_fecha_real.strftime('%d/%m/%Y') if primera_fecha_real else fecha_inicio.strftime('%d/%m/%Y'),  # CAMBIADO
                'fecha_fin': fecha_fin.strftime('%d/%m/%Y'),
                'total_entrenamientos': total_entrenamientos,
                'total_asistencias': total_asistencias,
                'porcentaje_asistencia': round(porcentaje_asistencia, 1),
                'total_objetivos_evaluados': total_objetivos_evaluados,
                'total_objetivos_cumplidos': total_objetivos_cumplidos,
                'porcentaje_objetivos': round(porcentaje_objetivos, 1),
            }
            
            # Generar el PDF
            pdf_path = generar_certificado_individual(datos_alumno, datos_reporte)
            
            # Verificar que el archivo existe
            if os.path.exists(pdf_path):
                # Abrir el archivo y devolverlo como respuesta
                pdf_file = open(pdf_path, 'rb')
                response = FileResponse(pdf_file, content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="certificado_{datos_alumno["id_persona"]}.pdf"'
                return response
            else:
                raise Http404("El archivo PDF no se pudo generar")
                
        except Matricula.DoesNotExist:
            raise Http404("Matrícula no encontrada")
        except Exception as e:
            raise Http404(f"Error al generar el certificado: {str(e)}")
    
    return redirect('reportes:index')


@login_required
def exportar_reporte_individual_excel(request, matricula_id):
    """Exporta el reporte individual a Excel"""
    if request.method == 'GET':
        fecha_inicio_str = request.GET.get('fechaInicio')
        fecha_fin_str = request.GET.get('fechaFin')
        
        try:
            # Convertir fechas
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            fecha_fin_ajustada = fecha_fin + timedelta(days=1)
            
            # Obtener la matrícula y datos del alumno
            matricula = Matricula.objects.select_related(
                'fk_alumno__fk_persona_alumno',
                'fk_categoria'
            ).get(idmatricula=matricula_id)
            
            persona = matricula.fk_alumno.fk_persona_alumno
            nombre_alumno = f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip()
            nombre_categoria = matricula.fk_categoria.nom_categoria
            
            # Obtener todas las asistencias del alumno en el rango de fechas
            asistencias = Asistencia.objects.filter(
                fk_matricula_ms=matricula,
                fecha_asistencia__gte=fecha_inicio,
                fecha_asistencia__lt=fecha_fin_ajustada
            ).select_related(
                'fk_sesion_ms__fk_entrenamiento'
            ).order_by('fecha_asistencia')
            
            # Crear el workbook de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Individual"
            
            # Estilos
            header_fill = PatternFill(start_color="1C417E", end_color="1C417E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            info_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            info_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws.merge_cells('A1:F1')
            ws['A1'] = 'REPORTE INDIVIDUAL DE OBJETIVOS POR ENTRENAMIENTO'
            ws['A1'].font = Font(bold=True, size=14, color="1C417E")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Información del alumno
            ws['A3'] = 'Alumno:'
            ws['B3'] = nombre_alumno
            ws['A3'].font = info_font
            ws['B3'].font = Font(size=11)
            
            ws['A4'] = 'Categoría:'
            ws['B4'] = nombre_categoria
            ws['A4'].font = info_font
            ws['B4'].font = Font(size=11)
            
            ws['A5'] = 'Período:'
            ws['B5'] = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
            ws['A5'].font = info_font
            ws['B5'].font = Font(size=11)
            
            # Aplicar color de fondo a la información
            for row in range(3, 6):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = info_fill
                    ws.cell(row=row, column=col).border = border
            
            # Encabezados de la tabla
            headers = ['Fecha', 'Asistencia', 'Tipo de Entrenamiento', 'Objetivo', 'Cumplido', 'Observaciones']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos de la tabla
            row_num = 8
            for asistencia in asistencias:
                # MODIFICADO: Obtener SOLO calificaciones de objetivos que fueron evaluados
                calificaciones = CalificacionObjetivos.objects.filter(
                    fk_asistencia=asistencia,
                    objetivo_evaluado=True  # Solo objetivos evaluados
                ).select_related('fk_asistencia')
                
                if calificaciones.exists():
                    for calificacion in calificaciones:
                        # Obtener el objetivo
                        try:
                            objetivo = Objetivos.objects.get(idobjetivos=calificacion.id_objetivo)
                            nombre_objetivo = objetivo.nom_objetivo
                        except Objetivos.DoesNotExist:
                            nombre_objetivo = "Objetivo no encontrado"
                        
                        # Fecha
                        ws.cell(row=row_num, column=1).value = asistencia.fecha_asistencia.strftime('%d/%m/%Y %H:%M')
                        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                        ws.cell(row=row_num, column=1).border = border
                        
                        # Asistencia
                        asistencia_texto = 'PRESENTE' if asistencia.asistencia == 1 else 'AUSENTE'
                        ws.cell(row=row_num, column=2).value = asistencia_texto
                        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
                        ws.cell(row=row_num, column=2).border = border
                        
                        # Color según asistencia
                        if asistencia_texto == 'PRESENTE':
                            ws.cell(row=row_num, column=2).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            ws.cell(row=row_num, column=2).font = Font(color="2E7D32", bold=True)
                        else:
                            ws.cell(row=row_num, column=2).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                            ws.cell(row=row_num, column=2).font = Font(color="C62828", bold=True)
                        
                        # Tipo de Entrenamiento
                        ws.cell(row=row_num, column=3).value = asistencia.fk_sesion_ms.fk_entrenamiento.nom_entrenamiento
                        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left')
                        ws.cell(row=row_num, column=3).border = border
                        
                        # Objetivo
                        ws.cell(row=row_num, column=4).value = nombre_objetivo
                        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', wrap_text=True)
                        ws.cell(row=row_num, column=4).border = border
                        
                        # MODIFICADO: Cumplido - Ya sabemos que objetivo_evaluado=True
                        cumplido_texto = 'SÍ' if calificacion.evaluacion else 'NO'
                        
                        ws.cell(row=row_num, column=5).value = cumplido_texto
                        ws.cell(row_num, column=5).alignment = Alignment(horizontal='center')
                        ws.cell(row_num, column=5).border = border
                        
                        # Color según cumplimiento
                        if cumplido_texto == 'SÍ':
                            ws.cell(row=row_num, column=5).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            ws.cell(row=row_num, column=5).font = Font(color="2E7D32", bold=True)
                        else:
                            ws.cell(row=row_num, column=5).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                            ws.cell(row=row_num, column=5).font = Font(color="C62828", bold=True)
                        
                        # Observaciones
                        ws.cell(row=row_num, column=6).value = calificacion.observaciones or asistencia.observaciones or '-'
                        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', wrap_text=True)
                        ws.cell(row=row_num, column=6).border = border
                        
                        row_num += 1
                # ELIMINADO: El bloque else que mostraba asistencias sin objetivos evaluados
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            
            # Ajustar altura de filas
            ws.row_dimensions[1].height = 25
            for row in range(7, row_num):
                ws.row_dimensions[row].height = 30
            
            # Preparar respuesta HTTP
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="reporte_individual_{persona.id_persona}_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx"'
            
            wb.save(response)
            return response
            
        except Matricula.DoesNotExist:
            raise Http404("Matrícula no encontrada")
        except Exception as e:
            raise Http404(f"Error al generar el archivo Excel: {str(e)}")
    
    return redirect('reportes:index')


@login_required
def enviar_reportes_por_correo(request, matricula_id):
    """Genera los reportes PDF y Excel y los envía por correo al alumno"""
    if request.method == 'GET':
        fecha_inicio_str = request.GET.get('fechaInicio')
        fecha_fin_str = request.GET.get('fechaFin')
        
        try:
            # Convertir fechas
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            fecha_fin_ajustada = fecha_fin + timedelta(days=1)
            
            # Obtener la matrícula y datos del alumno
            matricula = Matricula.objects.select_related(
                'fk_alumno__fk_persona_alumno',
                'fk_categoria'
            ).get(idmatricula=matricula_id)
            
            persona = matricula.fk_alumno.fk_persona_alumno
            nombre_alumno = f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip()
            email_alumno = persona.email_persona
            nombre_categoria = matricula.fk_categoria.nom_categoria
            
            # Obtener todas las asistencias del alumno en el rango de fechas
            asistencias = Asistencia.objects.filter(
                fk_matricula_ms=matricula,
                fecha_asistencia__gte=fecha_inicio,
                fecha_asistencia__lt=fecha_fin_ajustada
            ).select_related(
                'fk_sesion_ms__fk_entrenamiento'
            ).order_by('fecha_asistencia')
            
            # Obtener la primera fecha de asistencia real
            primera_fecha_real = None
            if asistencias.exists():
                primera_fecha_real = asistencias.first().fecha_asistencia
            
            # Calcular estadísticas para el PDF
            total_entrenamientos = asistencias.count()
            total_asistencias = asistencias.filter(asistencia=1).count()
            porcentaje_asistencia = (total_asistencias / total_entrenamientos * 100) if total_entrenamientos > 0 else 0
            
            calificaciones = CalificacionObjetivos.objects.filter(
                fk_asistencia__in=asistencias,
                objetivo_evaluado=True
            )
            
            total_objetivos_evaluados = calificaciones.count()
            total_objetivos_cumplidos = calificaciones.filter(evaluacion=True).count()
            porcentaje_objetivos = (total_objetivos_cumplidos / total_objetivos_evaluados * 100) if total_objetivos_evaluados > 0 else 0
            
            # Preparar datos del alumno para PDF
            datos_alumno = {
                'nombre_completo': nombre_alumno,
                'tipo_identificacion': persona.tipo_identidad,
                'numero_identificacion': persona.id_persona,
                'id_persona': persona.id_persona,
                'categoria': nombre_categoria,
            }
            
            # Preparar datos del reporte para PDF
            datos_reporte = {
                'fecha_inicio': primera_fecha_real.strftime('%d/%m/%Y') if primera_fecha_real else fecha_inicio.strftime('%d/%m/%Y'),
                'fecha_fin': fecha_fin.strftime('%d/%m/%Y'),
                'total_entrenamientos': total_entrenamientos,
                'total_asistencias': total_asistencias,
                'porcentaje_asistencia': round(porcentaje_asistencia, 1),
                'total_objetivos_evaluados': total_objetivos_evaluados,
                'total_objetivos_cumplidos': total_objetivos_cumplidos,
                'porcentaje_objetivos': round(porcentaje_objetivos, 1),
            }
            
            # ===== GENERAR PDF =====
            pdf_path = generar_certificado_individual(datos_alumno, datos_reporte)
            
            # ===== GENERAR EXCEL EN MEMORIA =====
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte Individual"
            
            # Estilos
            header_fill = PatternFill(start_color="1C417E", end_color="1C417E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            info_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            info_font = Font(bold=True, size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título principal
            ws.merge_cells('A1:F1')
            ws['A1'] = 'REPORTE INDIVIDUAL DE OBJETIVOS POR ENTRENAMIENTO'
            ws['A1'].font = Font(bold=True, size=14, color="1C417E")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Información del alumno
            ws['A3'] = 'Alumno:'
            ws['B3'] = nombre_alumno
            ws['A3'].font = info_font
            ws['B3'].font = Font(size=11)
            
            ws['A4'] = 'Categoría:'
            ws['B4'] = nombre_categoria
            ws['A4'].font = info_font
            ws['B4'].font = Font(size=11)
            
            ws['A5'] = 'Período:'
            ws['B5'] = f"{fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}"
            ws['A5'].font = info_font
            ws['B5'].font = Font(size=11)
            
            # Aplicar color de fondo a la información
            for row in range(3, 6):
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = info_fill
                    ws.cell(row=row, column=col).border = border
            
            # Encabezados de la tabla
            headers = ['Fecha', 'Asistencia', 'Tipo de Entrenamiento', 'Objetivo', 'Cumplido', 'Observaciones']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=7, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Datos de la tabla
            row_num = 8
            for asistencia in asistencias:
                calificaciones_asistencia = CalificacionObjetivos.objects.filter(
                    fk_asistencia=asistencia,
                    objetivo_evaluado=True
                ).select_related('fk_asistencia')
                
                if calificaciones_asistencia.exists():
                    for calificacion in calificaciones_asistencia:
                        try:
                            objetivo = Objetivos.objects.get(idobjetivos=calificacion.id_objetivo)
                            nombre_objetivo = objetivo.nom_objetivo
                        except Objetivos.DoesNotExist:
                            nombre_objetivo = "Objetivo no encontrado"
                        
                        # Fecha
                        ws.cell(row=row_num, column=1).value = asistencia.fecha_asistencia.strftime('%d/%m/%Y %H:%M')
                        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                        ws.cell(row=row_num, column=1).border = border
                        
                        # Asistencia
                        asistencia_texto = 'PRESENTE' if asistencia.asistencia == 1 else 'AUSENTE'
                        ws.cell(row=row_num, column=2).value = asistencia_texto
                        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')
                        ws.cell(row=row_num, column=2).border = border
                        
                        if asistencia_texto == 'PRESENTE':
                            ws.cell(row=row_num, column=2).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            ws.cell(row=row_num, column=2).font = Font(color="2E7D32", bold=True)
                        else:
                            ws.cell(row=row_num, column=2).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                            ws.cell(row=row_num, column=2).font = Font(color="C62828", bold=True)
                        
                        # Tipo de Entrenamiento
                        ws.cell(row=row_num, column=3).value = asistencia.fk_sesion_ms.fk_entrenamiento.nom_entrenamiento
                        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='left')
                        ws.cell(row=row_num, column=3).border = border
                        
                        # Objetivo
                        ws.cell(row=row_num, column=4).value = nombre_objetivo
                        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='left', wrap_text=True)
                        ws.cell(row=row_num, column=4).border = border
                        
                        # Cumplido
                        cumplido_texto = 'SÍ' if calificacion.evaluacion else 'NO'
                        ws.cell(row=row_num, column=5).value = cumplido_texto
                        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
                        ws.cell(row_num, column=5).border = border
                        
                        if cumplido_texto == 'SÍ':
                            ws.cell(row=row_num, column=5).fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
                            ws.cell(row=row_num, column=5).font = Font(color="2E7D32", bold=True)
                        else:
                            ws.cell(row=row_num, column=5).fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                            ws.cell(row=row_num, column=5).font = Font(color="C62828", bold=True)
                        
                        # Observaciones
                        ws.cell(row=row_num, column=6).value = calificacion.observaciones or asistencia.observaciones or '-'
                        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='left', wrap_text=True)
                        ws.cell(row=row_num, column=6).border = border
                        
                        row_num += 1
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 35
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 40
            
            # Ajustar altura de filas
            ws.row_dimensions[1].height = 25
            for row in range(7, row_num):
                ws.row_dimensions[row].height = 30
            
            # Guardar Excel en memoria
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            # ===== CREAR Y ENVIAR EMAIL =====
            asunto = f'Reporte de Objetivos y Asistencia - {nombre_alumno}'
            
            mensaje = f"""
            Estimado/a {nombre_alumno},
            
            Adjunto encontrará su reporte de objetivos y asistencia correspondiente al período del {fecha_inicio.strftime('%d/%m/%Y')} al {fecha_fin.strftime('%d/%m/%Y')}.
            
            El reporte incluye:
            - Certificado en formato PDF
            - Detalle completo de entrenamientos en formato Excel
            
            Categoría: {nombre_categoria}
            Total de entrenamientos: {total_entrenamientos}
            Asistencias: {total_asistencias} ({porcentaje_asistencia:.1f}%)
            Objetivos evaluados: {total_objetivos_evaluados}
            Objetivos cumplidos: {total_objetivos_cumplidos} ({porcentaje_objetivos:.1f}%)
            
            Saludos cordiales,
            Club Deportivo Atlético Santander
            """
            
            email = EmailMessage(
                asunto,
                mensaje,
                settings.DEFAULT_FROM_EMAIL,
                [email_alumno],
            )
            
            # Adjuntar PDF
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as pdf_file:
                    email.attach(
                        f'certificado_{persona.id_persona}.pdf',
                        pdf_file.read(),
                        'application/pdf'
                    )
            
            # Adjuntar Excel
            email.attach(
                f'reporte_individual_{persona.id_persona}_{fecha_inicio.strftime("%Y%m%d")}_{fecha_fin.strftime("%Y%m%d")}.xlsx',
                excel_buffer.getvalue(),
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Enviar email
            email.send()
            
            # Limpiar archivo PDF temporal
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
            
            # Redirigir con mensaje de éxito
            from django.contrib import messages
            messages.success(request, f'Reportes enviados exitosamente a {email_alumno}')
            
            # Redirigir de vuelta al reporte individual con los mismos parámetros
            return redirect(f'/reportes/generar-individual/?success=email_enviado&categoria={matricula.fk_categoria_id}&idAlumno={matricula_id}&fechaInicio={fecha_inicio_str}&fechaFin={fecha_fin_str}')
            
        except Matricula.DoesNotExist:
            raise Http404("Matrícula no encontrada")
        except Exception as e:
            from django.contrib import messages
            messages.error(request, f'Error al enviar el correo: {str(e)}')
            return redirect('reportes:index')
    
    return redirect('reportes:index')


def listar_entrenamientos(request):
    """Vista para listar todos los entrenamientos con filtros"""
    categorias = Categoria.objects.all().order_by('idcategoria')
    
    # Obtener parámetros de filtro
    categoria_id = request.GET.get('categoria', '')
    fecha_inicio_str = request.GET.get('fechaInicio', '')
    fecha_fin_str = request.GET.get('fechaFin', '')
    
    # NUEVO: Obtener nombre de categoría si existe
    nombre_categoria_filtro = None
    if categoria_id:
        try:
            nombre_categoria_filtro = Categoria.objects.get(idcategoria=categoria_id).nom_categoria
        except Categoria.DoesNotExist:
            pass
    
    # Iniciar query de sesiones
    sesiones_query = Sesionentrenamiento.objects.select_related(
        'fk_entrenamiento'
    ).prefetch_related(
        'matriculas__fk_categoria'
    )
    
    # Aplicar filtros
    if categoria_id:
        # Filtrar por categoría a través de matrículas
        sesiones_query = sesiones_query.filter(
            matriculas__fk_categoria_id=categoria_id
        ).distinct()
    
    if fecha_inicio_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            sesiones_query = sesiones_query.filter(
                fecha_entrenamiento__gte=fecha_inicio
            )
        except ValueError:
            pass
    
    if fecha_fin_str:
        try:
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            fecha_fin_ajustada = fecha_fin + timedelta(days=1)
            sesiones_query = sesiones_query.filter(
                fecha_entrenamiento__lt=fecha_fin_ajustada
            )
        except ValueError:
            pass
    
    # Ordenar por fecha descendente
    sesiones = sesiones_query.order_by('-fecha_entrenamiento')
    
    entrenamientos_data = []
    
    for sesion in sesiones:
        # Obtener las categorías únicas de esta sesión
        if categoria_id:
            # Si hay filtro de categoría, mostrar solo esa categoría
            categorias_sesion = [
                Categoria.objects.get(idcategoria=categoria_id).nom_categoria
            ]
        else:
            # Mostrar todas las categorías de la sesión
            categorias_sesion = sesion.matriculas.values_list(
                'fk_categoria__nom_categoria', 
                flat=True
            ).distinct()
        
        # Contar matrículas según filtro de categoría
        if categoria_id:
            total_inscritos = sesion.matriculas.filter(
                fk_categoria_id=categoria_id
            ).count()
        else:
            total_inscritos = sesion.matriculas.count()
        
        entrenamientos_data.append({
            'idsesion': sesion.idsesion,
            'fecha': sesion.fecha_entrenamiento,
            'categorias': ', '.join(categorias_sesion) if categorias_sesion else 'Sin categoría',
            'nombre_entrenamiento': sesion.fk_entrenamiento.nom_entrenamiento,
            'total_inscritos': total_inscritos
        })
    
    context = {
        'categorias': categorias,
        'entrenamientos': entrenamientos_data,
        'tipoReporte': None,
        'nombre_categoria_filtro': nombre_categoria_filtro,  # NUEVO
    }
    
    return render(request, 'reportes.html', context)


@require_http_methods(["GET"])
def obtener_detalle_entrenamiento(request, sesion_id):
    """API para obtener el detalle de un entrenamiento específico"""
    try:
        sesion = Sesionentrenamiento.objects.select_related(
            'fk_entrenamiento'
        ).prefetch_related(
            'matriculas__fk_alumno__fk_persona_alumno',
            'matriculas__fk_categoria'
        ).get(idsesion=sesion_id)
        
        # Obtener todas las asistencias de esta sesión
        asistencias = Asistencia.objects.filter(
            fk_sesion_ms=sesion
        ).select_related(
            'fk_matricula_ms__fk_alumno__fk_persona_alumno'
        ).order_by('fk_matricula_ms__fk_alumno__fk_persona_alumno__ape1_persona')
        
        # Obtener objetivos del entrenamiento
        objetivos_entrenamiento = EntrenamientoObjetivo.objects.filter(
            fk_entrenamiento=sesion.fk_entrenamiento
        ).select_related('fk_objetivo').values_list('fk_objetivo__nom_objetivo', flat=True)
        
        alumnos_detalle = []
        
        for asistencia in asistencias:
            persona = asistencia.fk_matricula_ms.fk_alumno.fk_persona_alumno
            nombre_completo = f"{persona.ape1_persona} {persona.ape2_persona or ''} {persona.nom1_persona} {persona.nom2_persona or ''}".strip()
            
            # Obtener calificaciones de objetivos para este alumno en esta sesión
            calificaciones = CalificacionObjetivos.objects.filter(
                fk_asistencia=asistencia,
                objetivo_evaluado=True
            )
            
            objetivos_alumno = []
            for calificacion in calificaciones:
                try:
                    objetivo = Objetivos.objects.get(idobjetivos=calificacion.id_objetivo)
                    objetivos_alumno.append({
                        'nombre': objetivo.nom_objetivo,
                        'cumplido': calificacion.evaluacion,
                        'observaciones': calificacion.observaciones or 'Sin observaciones'
                    })
                except Objetivos.DoesNotExist:
                    pass
            
            alumnos_detalle.append({
                'nombre_completo': nombre_completo,
                'asistio': asistencia.asistencia == 1,
                'categoria': asistencia.fk_matricula_ms.fk_categoria.nom_categoria,
                'objetivos': objetivos_alumno,
                'observaciones_generales': asistencia.observaciones or 'Sin observaciones'
            })
        
        return JsonResponse({
            'success': True,
            'sesion': {
                'idsesion': sesion.idsesion,
                'fecha': sesion.fecha_entrenamiento.strftime('%d/%m/%Y %H:%M'),
                'nombre_entrenamiento': sesion.fk_entrenamiento.nom_entrenamiento,
                'descripcion': sesion.fk_entrenamiento.descripcion,
                'objetivos_entrenamiento': list(objetivos_entrenamiento),
                'total_inscritos': asistencias.count(),
                'total_asistieron': asistencias.filter(asistencia=1).count()
            },
            'alumnos': alumnos_detalle
        })
        
    except Sesionentrenamiento.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Sesión de entrenamiento no encontrada'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

